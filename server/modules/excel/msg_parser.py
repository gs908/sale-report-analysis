
from openpyxl import load_workbook
from typing import Generator, Iterator
import re


# ─────────────────────────────────────────────
# Group message parsing — openpyxl 只读流式
# ─────────────────────────────────────────────
def is_auto_sender(sender: str) -> bool:
    """判断是否为自动消息发送人（系统/机器人）"""
    if not sender:
        return False
    auto_patterns = ["系统", "机器人", "小助手", "Auto", "System", "admin", "Admin"]
    return any(p in sender for p in auto_patterns)


def parse_group_messages(
    file_path: str,
    sheet_name: str = "Sheet1",
    batch_size: int = 5000,
    truncate: int = 500,
) -> Generator[dict, None, None]:
    """
    openpyxl 只读模式流式遍历群消息 Excel，按 group_id 分组
    Yields: 每个 group 的一条已处理数据
    """
    time_fmt = "%Y-%m-%d %H:%M:%S"

    wb = load_workbook(filename=file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    headers: list[str] = []
    groups_map: dict = {}
    groups_buffer_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue

        row_dict = {headers[j]: row[j] for j in range(len(headers))}
        group_id = row_dict.get("群ID")
        if not group_id or str(group_id).strip() == "":
            continue
        group_id = str(group_id).strip()

        if group_id not in groups_map:
            groups_map[group_id] = {
                "group_id": group_id,
                "group_name": row_dict.get("接收人或群", ""),
                "msgs": [],
            }
        groups_map[group_id]["msgs"].append(row_dict)
        groups_buffer_count += 1

        # 每累积 batch_size 条消息，输出一批已完成的 group
        if groups_buffer_count >= batch_size:
            for gid, info in groups_map.items():
                yield _process_group(info, truncate, time_fmt)
            groups_map.clear()
            groups_buffer_count = 0

    wb.close()

    # 剩余 group
    for gid, info in groups_map.items():
        yield _process_group(info, truncate, time_fmt)


def _minutes_to_sec(minutes: float | None) -> int:
    """分钟数转秒数"""
    if minutes is None:
        return 0
    return int(round(minutes * 60))


def _sec_to_str(total_sec: int) -> str:
    """秒数格式化为'X小时X分X秒'"""
    if total_sec == 0:
        return "0秒"
    parts = []
    hours = total_sec // 3600
    mins = (total_sec % 3600) // 60
    secs = total_sec % 60
    if hours > 0:
        parts.append(f"{hours}小时")
    if mins > 0:
        parts.append(f"{mins}分")
    if secs > 0 or not parts:
        parts.append(f"{secs}秒")
    return "".join(parts)


def _process_group(info: dict, truncate: int, time_fmt: str) -> dict:
    """处理单个 group，计算统计指标"""
    sorted_msgs = sorted(
        info["msgs"],
        key=lambda x: str(x.get("消息时间") or ""),
    )

    processed_msgs: list[dict] = []
    time_intervals: list[float] = []
    last_time = None  # 最后一条非自动消息的时间
    start_time_dt = None  # 第一条消息（自动建群）时间
    first_real_msg_time = None  # 第一条非自动消息时间
    # 从 group_name（接收人或群字段）中提取 issue_id
    issue_id_match = re.search(r"WT-\d+", info.get("group_name", ""), re.IGNORECASE)
    issue_id = issue_id_match.group(0).upper() if issue_id_match else "Unknown"
    has_real_message = False  # 是否有真实用户消息

    for msg in sorted_msgs:
        raw_time_str = msg.get("消息时间")
        curr_time_str = str(raw_time_str) if raw_time_str else ""
        curr_content = str(msg.get("消息内容") or "")
        sender = str(msg.get("发送人姓名") or "")
        is_auto_msg = is_auto_sender(sender)

        # 记录第一条消息时间（自动建群时间）
        if curr_time_str and start_time_dt is None:
            try:
                start_time_dt = datetime.strptime(curr_time_str, time_fmt)
            except Exception:
                pass

        # 计算响应时间时，屏蔽自动消息
        if curr_time_str and not is_auto_msg:
            has_real_message = True
            try:
                curr_dt = datetime.strptime(curr_time_str, time_fmt)
                if first_real_msg_time is None:
                    first_real_msg_time = curr_dt
                if last_time is not None:
                    diff = (curr_dt - last_time).total_seconds() / 60
                    time_intervals.append(diff)
                last_time = curr_dt
            except Exception:
                pass

        # 所有消息都保留在 chat_msgs 中
        processed_msgs.append({
            "time": curr_time_str,
            "sender": sender,
            "text": curr_content,
        })

    # 首次响应秒数
    first_resp_sec = 0
    if start_time_dt and first_real_msg_time:
        first_resp_sec = int(round((first_real_msg_time - start_time_dt).total_seconds()))

    # 间隔秒数
    max_gap_sec = 0
    avg_gap_sec = 0
    if time_intervals:
        max_gap_sec = int(round(max(time_intervals) * 60))
        avg_gap_sec = int(round(sum(time_intervals) / len(time_intervals) * 60))

    # 问题时长（秒）
    duration_sec = 0
    if start_time_dt and last_time:
        duration_sec = int(round((last_time - start_time_dt).total_seconds()))

    first_msg_time = processed_msgs[0]["time"] if processed_msgs else None
    last_msg_time = processed_msgs[-1]["time"] if processed_msgs else None

    real_msg_count = sum(1 for msg in processed_msgs if not is_auto_sender(msg["sender"]))

    return {
        "issue_id": issue_id,
        "group_id": info.get("group_id", ""),
        "group_name": info["group_name"],
        "created_at": first_msg_time[:10] if first_msg_time else None,
        "statistics": {
            "msg_count": len(processed_msgs),
            "real_msg_count": real_msg_count,
            "first_resp_sec": first_resp_sec,
            "first_resp_str": _sec_to_str(first_resp_sec),
            "max_gap_sec": max_gap_sec,
            "max_gap_str": _sec_to_str(max_gap_sec),
            "avg_gap_sec": avg_gap_sec,
            "avg_gap_str": _sec_to_str(avg_gap_sec),
            "duration_sec": duration_sec,
            "duration_str": _sec_to_str(duration_sec),
            "first_msg_time": first_msg_time,
            "last_msg_time": last_msg_time,
            "only_auto_msg": not has_real_message,
            "less_than_3": real_msg_count < 3,
        },
    }


def parse_group_messages_flat(
    file_path: str,
    sheet_name: str = "Sheet1",
    batch_size: int = 5000,
    truncate: int = 500,
) -> Iterator[dict]:
    for group_doc in parse_group_messages(file_path, sheet_name, batch_size, truncate):
        yield group_doc


# ─────────────────────────────────────────────
# Raw Group message parsing — 原始记录直接入库
# ─────────────────────────────────────────────
def parse_raw_group_messages(
    file_path: str,
    sheet_name: str = "Sheet1",
    batch_size: int = 5000,
) -> Generator[list[dict], None, None]:
    """
    读取群消息 Excel，按行yield，每行作为一条独立记录
    不做任何处理，直接写入 raw_groups 集合
    """
    wb = load_workbook(filename=file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    headers: list[str] = []
    batch: list[dict] = []
    total_rows = 0
    skipped_rows = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue

        total_rows += 1
        if total_rows % 10000 == 0:
            print(f"[Step2-raw] 已读取 {total_rows} 行 ...")

        row_dict = {headers[j]: row[j] for j in range(len(headers))}

        # 只导入群消息（排除非群消息行）
        is_group = row_dict.get("是否群消息", "")
        if str(is_group).strip() != "是":
            skipped_rows += 1
            continue

        # 从"接收人或群"字段中提取 issue_id（WT- + 5位数字）
        receiver = str(row_dict.get("接收人或群") or "").strip()
        issue_id_match = re.search(r"WT-\d+", receiver, re.IGNORECASE)
        issue_id = issue_id_match.group(0).upper() if issue_id_match else None

        batch.append({
            "issue_id": issue_id,
            "group_id": str(row_dict.get("群ID") or "").strip(),
            "sender_id": str(row_dict.get("发送ID") or "").strip(),
            "sender": str(row_dict.get("发送人姓名") or "").strip(),
            "msg_type": str(row_dict.get("消息类型") or "").strip(),
            "receiver_id": str(row_dict.get("接收人ID") or "").strip(),
            "receiver": receiver,
            "msg_time": str(row_dict.get("消息时间") or "").strip(),
            "msg_content": str(row_dict.get("消息内容") or "").strip(),
            "msg_id": row_dict.get("ID"),  # 消息原生ID，每行唯一
        })

        if len(batch) >= batch_size:
            yield batch
            batch = []

    if batch:
        yield batch

    wb.close()
    print(f"[Step2-raw] 读取完成：共 {total_rows} 行，其中跳过（非群消息）{skipped_rows} 行")


def parse_raw_group_messages_flat(
    file_path: str,
    sheet_name: str = "Sheet1",
    batch_size: int = 5000,
) -> Iterator[dict]:
    """展平版本，每次 yield 单条记录"""
    for batch in parse_raw_group_messages(file_path, sheet_name, batch_size):
        for doc in batch:
            yield doc